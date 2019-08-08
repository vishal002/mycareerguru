from openpyxl import load_workbook
import MySQLdb

wb = load_workbook(filename='./Sample Courses data for Aditya.xlsx')
ws = wb['Sheet1'] # ws is now an IterableWorksheet

db = MySQLdb.connect("localhost", "root", "root", "mcgdb")
cursor = db.cursor()

count = 0
error_count = 0

def parse_text(text):
	return text.replace("\"", " ").strip()

def get_major_field(field):
	query = "select id from majorfields_majorfields where name=\"%s\"" %(field)
	cursor.execute(query)
	data = cursor.fetchall()
	return data[0][0] if len(data) else None

for index, row in enumerate(ws.iter_rows(min_row=2 , max_col=17)):
	try:
		if not row[0].value:
			break
		name 			= parse_text(row[0].value) if row[0].value else ""
		description		= parse_text(row[1].value) if row[1].value else ""
		duration 		= parse_text(row[2].value) if row[2].value else ""
		level 			= parse_text(row[3].value) if row[3].value else ""
		course_type 	= parse_text(row[4].value) if row[4].value else ""
		eligibility 	= parse_text(row[5].value) if row[5].value else ""
		overview 		= parse_text(row[6].value) if row[6].value else ""
		eligibility_2 	= parse_text(row[7].value) if row[7].value else ""
		syllabus 		= parse_text(row[8].value) if row[8].value else ""
		suitability 	= parse_text(row[10].value) if row[10].value else ""
		beneficial 		= parse_text(row[11].value) if row[11].value else ""
		employment_areas= parse_text(row[12].value) if row[12].value else ""
		job_types 		= parse_text(row[13].value) if row[13].value else ""
		advance_course 	= parse_text(row[14].value) if row[14].value else ""
		related_course 	= parse_text(row[15].value) if row[15].value else ""
		field 			= parse_text(row[16].value) if row[16].value else ""
		careers 		= ""
		recuiting_companies= ""
		other_details 	= ""

		major_field_id = str(get_major_field(field))
		if not major_field_id:
			raise Exception('Major field not found in database.')

		print major_field_id

		entry = ("\""+name+"\",\""+description+"\",\""+duration+"\",\""+level+
					"\",\""+course_type+"\",\""+eligibility+"\",\""+overview+
					"\",\""+eligibility_2+"\",\""+syllabus+"\",\""+suitability+
					"\",\""+beneficial+"\",\""+employment_areas+"\",\""+job_types+
					"\",\""+advance_course+"\",\""+related_course+"\",\""+careers+
					"\",\""+recuiting_companies+"\",\""+other_details+"\",\""+major_field_id+"\"")

		query = ("insert into courses_course (name, description, duration, level,"+
		 			"course_type, eligibility, overview, eligibility_2, syllabus,"+
		 			" suitability, beneficial, employment_areas, job_types,"+
		 			" advance_course, related_course, careers, recuiting_companies, other_details, majorfield_id)"+
		 			" values ("+entry.encode('utf-8')+");")

		#print query
		cursor.execute(query)
		data = cursor.fetchall()
		db.commit()
		count += 1
	except Exception, fault:
		print fault
		error_count += 1
		print "Could not parse row : "+str(index)

print "Rows parsed : "+ str(count)
print "Rows errored : "+ str(error_count)

# Course Title -- 0
# Description -- 1
# Duration -- 2
# Level -- 3
# Type -- 4
# Eligibility -- 5
# Overview -- 6
# Eligibility 2 -- 7
# Syllabus -- 8
# Colleges -- 9
# Suitability -- 10
# Beneficial -- 11
# Employment Areas -- 12
# Job Types -- 13
# Advance Courses -- 14
# Related Courses -- 15