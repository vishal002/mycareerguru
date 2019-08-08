from openpyxl import load_workbook
import MySQLdb

wb = load_workbook(filename='./Sample Entrance Exam data for Aditya.xlsx')
ws = wb['Sheet1'] # ws is now an IterableWorksheet

db = MySQLdb.connect("localhost", "root", "root", "mcgdb")
cursor = db.cursor()

count = 0
error_count = 0


def parse_text(text):
	return text.replace("\"", " ").strip()

for index, row in enumerate(ws.iter_rows(min_row=2, max_col=25)):
	try:
		if not row[0].value:
			break
		name 				= parse_text(row[0].value) if row[0].value else ""
		about 				= parse_text(row[1].value) if row[1].value else ""
		eligibility 		= parse_text(row[2].value) if row[2].value else ""
		age_limit 			= parse_text(row[3].value) if row[3].value else ""
		marital_status 		= parse_text(row[4].value) if row[4].value else ""
		physical_standard 	= parse_text(row[5].value) if row[5].value else ""
		syllabus 		  	= parse_text(row[6].value) if row[6].value else ""
		exam_pattern  		= parse_text(row[7].value) if row[7].value else ""
		validity 		  	= parse_text(row[8].value) if row[8].value else ""
		application_procedure= parse_text(row[9].value) if row[9].value else ""
		application_form  	= parse_text(row[10].value) if row[10].value else ""
		application_fee 	= parse_text(row[11].value) if row[11].value else ""
		exam_date 			= parse_text(row[12].value) if row[12].value else ""
		exam_centres 		= parse_text(row[13].value) if row[13].value else ""
		
		addr_line1 	= parse_text(row[14].value) if row[14].value else ""
		addr_line2 	= parse_text(row[15].value) if row[15].value else ""
		addr_line3 	= parse_text(row[16].value) if row[16].value else ""
		addr_line4 	= parse_text(row[17].value) if row[17].value else ""
		addr_line5 	= parse_text(row[18].value) if row[18].value else ""
		addr_line6 	= parse_text(row[19].value) if row[19].value else ""

		full_address = addr_line1+" "+addr_line2+" "+addr_line3+" "+addr_line4+" "+addr_line5+" "+addr_line6
		
		phone 				= parse_text(row[20].value) if row[20].value else ""
		website 			= parse_text(row[21].value) if row[21].value else ""
		fax 				= parse_text(row[22].value) if row[22].value else ""
		email 				= parse_text(row[23].value) if row[23].value else ""
		level	 			= parse_text(row[24].value) if row[23].value else ""

		other_details 		= ""
		field				= ""
		preparation_tips    = ""
		exam_results		= ""
		title           	= ""
		latest_development 	= ""
		previous_years_question_papers = ""
		notification    	= ""
		important_dates 	= ""
		study_material  	= ""
		admit_card      	= ""
		exam_toppers    	= ""
		subject_experts 	= ""
		exam_cut_offs   	= ""
		preparation_videos 	= ""
		documents_to_be_carried = ""
		acknowledgement_of_form = ""
		score_evalutaion   	= ""
		faqs               	= ""

		field = "Engineering"
		query = "select id from majorfields_majorfields where name=\"%s\""%(field.strip())
		cursor.execute(query)

		data = cursor.fetchall()
		field_id=0
		if len(data):
			# field is already present in list
			field_id = str(data[0][0])
		else:
			msg = "field does not exists -- " %str(field)
			print msg
			raise Exception(msg) 
		# insert exam data
		entry = ("\""+name+"\",\""+about+"\",\""+eligibility+"\",\""+age_limit+
					"\",\""+marital_status+"\",\""+physical_standard+"\",\""+syllabus+
					"\",\""+exam_pattern+"\",\""+validity+"\",\""+application_procedure+
					"\",\""+application_form+"\",\""+application_fee+"\",\""+exam_date+
					"\",\""+exam_centres+"\",\""+full_address+"\",\""+phone+
					"\",\""+website+"\",\""+fax+"\",\""+email+
					"\",\""+other_details+"\",\""+field_id+"\",\""+level+
					"\",\""+preparation_tips+"\",\""+exam_results+
					"\",\""+title+"\",\""+latest_development+
					"\",\""+previous_years_question_papers+"\",\""+notification+
					"\",\""+important_dates+"\",\""+study_material+"\",\""+admit_card+
					"\",\""+exam_toppers+"\",\""+subject_experts+"\",\""+exam_cut_offs+
					"\",\""+preparation_videos+"\",\""+documents_to_be_carried+
					"\",\""+acknowledgement_of_form+"\",\""+score_evalutaion+
					"\",\""+faqs+"\"")

		query = ("insert into exam_exam (name, about, eligibility, age_limit,"+
		 			"marital_status, physical_standard, syllabus, exam_pattern, validity,"+
		 			" application_procedure, application_form, application_fee, exam_date,"+
		 			" exam_centres, full_address, phone, website, fax, email,"+
		 			" other_details, field_id, level, preparation_tips, exam_results,"+
		 			" title, latest_development, previous_years_question_papers, notification,"+
		 			" important_dates, study_material, admit_card, exam_toppers, subject_experts,"+
		 			" exam_cut_offs, preparation_videos, documents_to_be_carried,"+
		 			" acknowledgement_of_form, score_evalutaion, faqs"+
		 			") values ("+entry.encode('utf-8')+");")
		
		cursor.execute(query)
		data = cursor.fetchall()

		db.commit()
		count += 1
	except Exception, fault:
		print "------------------------- "
		error_count += 1
		print "Could not parse row : "+str(index)+ " | Error : ", fault
		print "------------------------- "

print "Rows parsed : "+ str(count)
print "Rows errored : "+ str(error_count)
