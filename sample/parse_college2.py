from openpyxl import load_workbook
import MySQLdb

wb = load_workbook(filename='/var/www/mycareerguru/mycareerguru/mycareerguru/'
						'sample/Sample CollegesV2 data for Aditya.xlsx')
ws = wb['Sheet1'] # ws is now an IterableWorksheet

db = MySQLdb.connect("localhost", "root", "root", "mcgdb")
cursor = db.cursor()

count = 0
error_count = 0

def parse_text(text):
	return text.replace("\"", " ").strip()

for index, row in enumerate(ws.iter_rows(min_row=2, max_col=23)):
	try:
		if not row[0].value:
			break
		name 			= parse_text(row[0].value) if row[0].value else ""
		city 			= parse_text(row[1].value) if row[1].value else ""
		state 			= parse_text(row[2].value) if row[2].value else ""
		establishment 	= parse_text(row[3].value) if row[3].value else ""
		accreditted_by 	= parse_text(row[4].value) if row[4].value else ""
		institute_type 	= parse_text(row[5].value) if row[5].value else ""
		affiliated_to 	= parse_text(row[6].value) if row[6].value else ""
		rank_index 		= parse_text(row[8].value) if row[8].value else "" 
		about 			= parse_text(row[9].value) if row[9].value else ""
		visiting_companies = parse_text(row[10].value) if row[10].value else ""
		placements 		= parse_text(row[13].value) if row[13].value else ""
		facilities 		= parse_text(row[14].value) if row[14].value else ""
		faculty 		= parse_text(row[15].value) if row[15].value else ""
		scholarship_and_finance = parse_text(row[16].value) if row[16].value else ""
		pin_code 		= parse_text(row[17].value) if row[17].value else ""
		email 			= parse_text(row[18].value) if row[18].value else ""
		website 		= parse_text(row[19].value) if row[19].value else ""
		phone 			= parse_text(row[20].value) if row[20].value else ""
		full_address 	= parse_text(row[21].value) if row[21].value else ""
		avg_salary	 	= parse_text(row[22].value) if row[22].value else ""
		fees            = parse_text("")
		rating 			= parse_text("")
		other_details   = parse_text("")
		naac_rating		= parse_text("")
		country			= parse_text("")
		fax				= parse_text("")

		entry = ("\""+name+"\",\""+city+"\",\""+state+"\",\""+establishment+
					"\",\""+accreditted_by+"\",\""+institute_type+"\",\""+affiliated_to+
					"\",\""+rank_index+"\",\""+about+"\",\""+visiting_companies+
					"\",\""+placements+"\",\""+facilities+"\",\""+faculty+
					"\",\""+scholarship_and_finance+"\",\""+pin_code+"\",\""+email+
					"\",\""+website+"\",\""+phone+"\",\""+full_address+
					"\",\""+avg_salary+"\",\""+fees+"\",\""+rating+
					"\",\""+other_details+"\",\""+naac_rating+"\",\""+country+"\",\""+fax+"\"")
		query = ("insert into colleges_college (name, city, state, establishment,"+
		 			"accreditted_by, institute_type, affiliated_to, rank_index, about,"+
		 			" visiting_companies, placements, facilities, faculty,"+
		 			" scholarship_and_finance, pin_code, email, website, phone,"+
		 			" full_address, avg_salary, fees, rating, other_details, naac_rating, country, fax) values ("+
		 			entry.encode('utf-8')+");")
		#print query
		cursor.execute(query)
		data = cursor.fetchall()
		db.commit()
		count += 1
	except Exception, fault:
		print fault
		error_count += 1
		print "Could not parse row : "+str(index)

print "Rows parsed : "+ str(count)
print "Rows errored : "+ str(error_count)



# 0 -- Name of the Institution
# 1 -- City
# 2 -- State
# 3 -- Established in
# 4 -- Accredited By
# 5 -- Type
# 6 -- Affiliated To
# 7 -- Facilities
# 8 -- Rank Index
# 9 -- About
# 10 -- Visiting Companies
# 11 -- # of Courses
# 12 -- Courses
# 13 -- Placements
# 14 -- Available College Facilities
# 15 -- Faculty
# 16 -- Scholarship and Finance
# 17 -- PIN
# 18 -- Email
# 19 -- Website
# 20 -- Phone
# 21 -- Address
# 22 -- Average Salary