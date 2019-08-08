from openpyxl import load_workbook
import MySQLdb

wb = load_workbook(filename='./Sample Data of 1000 Schools.xlsx.xlsx')
ws = wb['Sheet3'] # ws is now an IterableWorksheet

db = MySQLdb.connect("localhost", "root", "root", "mcgdb")
cursor = db.cursor()

count = 0
error_count = 0


def parse_text(text):
	return text.replace("\"", " ").strip()

for index, row in enumerate(ws.iter_rows(min_row=2, max_col=24)):
	try:
		if not row[0].value:
			break

		name 		= parse_text(row[0].value) if row[0].value else ""
		full_address= parse_text(row[1].value) if row[1].value else ""
		city		= parse_text(row[2].value) if row[2].value else ""
		state 		= parse_text(row[3].value) if row[3].value else ""
		pincode 	= parse_text(row[4].value) if row[4].value else ""
		country 	= parse_text(row[5].value) if row[5].value else ""
		phone 		= parse_text(row[6].value) if row[6].value else ""
		mobile  	= parse_text(row[7].value) if row[7].value else ""
		fax 		= parse_text(row[8].value) if row[8].value else ""
		website 	= parse_text(row[9].value) if row[9].value else ""
		class_info  = parse_text(row[10].value) if row[10].value else ""
		board 		= parse_text(row[11].value) if row[11].value else ""
		info_3 		= parse_text(row[12].value) if row[12].value else ""
		info_4 		= parse_text(row[13].value) if row[13].value else ""
		info_5 		= parse_text(row[14].value) if row[14].value else ""
		
		# insert exam data
		entry = ("\""+name+"\",\""+full_address+"\",\""+city+"\",\""+state+
					"\",\""+pincode+"\",\""+country+"\",\""+phone+
					"\",\""+mobile+"\",\""+fax+"\",\""+website+
					"\",\""+class_info+"\",\""+board+"\",\""+info_3+
					"\",\""+info_4+"\",\""+info_5+"\"")
		query = ("insert into school_school (name, full_address, city, state,"+
		 			"pin_code, country, phone, mobile, fax,"+
		 			" website, class_info, board, info_3,"+
		 			" info_4, info_5"+
		 			") values ("+entry.encode('utf-8')+");")
		cursor.execute(query)
		data = cursor.fetchall()
		db.commit()
		count += 1
	except Exception, fault:
		print "------------------------- "
		error_count += 1
		print "Could not parse row : "+str(index)+ " | Error : ", fault
		print "------------------------- "

print "Rows parsed : "+ str(count)
print "Rows errored : "+ str(error_count)