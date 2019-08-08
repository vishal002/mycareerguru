from openpyxl import load_workbook
import MySQLdb
import query_helper

wb = load_workbook(filename='./Sample Data of Colleges.xlsx')
ws = wb['Sheet1'] # ws is now an IterableWorksheet

db = MySQLdb.connect("localhost", "root", "root", "mcgdb")
cursor = db.cursor()

count = 0
error_count = 0


def parse_text(text):
	return text.replace("\"", " ").strip()

for index, row in enumerate(ws.iter_rows(min_row=2, max_col=23)):
	try:
		if not row[0].value:
			break
		name 			= parse_text(row[0].value) if row[0].value else ""
		full_address 	= parse_text(row[1].value) if row[1].value else ""
		city 			= parse_text(row[2].value) if row[2].value else ""
		state 			= parse_text(row[3].value) if row[3].value else ""
		pin_code 		= parse_text(row[4].value) if row[4].value else ""
		country 		= parse_text(row[5].value) if row[5].value else ""
		phone 			= parse_text(row[6].value) if row[6].value else ""
		fax  			= parse_text(row[8].value) if row[8].value else ""
		website 		= parse_text(row[9].value) if row[9].value else ""
		affiliated_to 	= parse_text(row[10].value) if row[10].value else ""
		accreditted_by 	= parse_text(row[11].value) if row[11].value else ""

		establishment 	= parse_text("")
		institute_type 	= parse_text("")
		rank_index 		= parse_text("")
		about 			= parse_text("")
		visiting_companies = parse_text("")
		placements 		= parse_text("")
		facilities 		= parse_text("")
		faculty 		= parse_text("")
		scholarship_and_finance = parse_text("")
		email 			= parse_text("")
		avg_salary	 	= parse_text("")
		fees            = parse_text("")
		rating 			= parse_text("")
		other_details   = parse_text("")
		naac_rating		= parse_text("")

		query = query_helper.create_college_query((name, city, state, establishment, accreditted_by, institute_type, affiliated_to,
												   rank_index, about, visiting_companies, placements, facilities, faculty,
												   scholarship_and_finance, pin_code, email, website, phone, full_address,
												   avg_salary, fees, rating, other_details, naac_rating, country, fax))

		'''
		entry = ("\""+name+"\",\""+city+"\",\""+state+"\",\""+establishment+
					"\",\""+accreditted_by+"\",\""+institute_type+"\",\""+affiliated_to+
					"\",\""+rank_index+"\",\""+about+"\",\""+visiting_companies+
					"\",\""+placements+"\",\""+facilities+"\",\""+faculty+
					"\",\""+scholarship_and_finance+"\",\""+pin_code+"\",\""+email+
					"\",\""+website+"\",\""+phone+"\",\""+full_address+
					"\",\""+avg_salary+"\",\""+fees+"\",\""+rating+
					"\",\""+other_details+"\",\""+naac_rating+"\",\""+country+"\",\""+fax+"\"")
		query = ("insert into colleges_college (name, city, state, establishment,"+
		 			"accreditted_by, institute_type, affiliated_to, rank_index, about,"+
		 			" visiting_companies, placements, facilities, faculty,"+
		 			" scholarship_and_finance, pin_code, email, website, phone,"+
		 			" full_address, avg_salary, fees, rating, other_details, naac_rating, country, fax) values ("+
		 			entry.encode('utf-8')+");")
		'''

		cursor.execute(query)
		data = cursor.fetchall()

		# get id of college just inserted
		query = "select id from colleges_college where name=\"%s\""%(name)
		cursor.execute(query)
		data = cursor.fetchall()
		college_id = data[0][0]

		#print "college_id is ", college_id
		# courses list in row 12, parse and put in college course mapping table
		courses = row[12].value.split(",")

		for l in courses:
			try:
				field, course, duration = l.split("-")
			except ValueError, fault:
				print "try 1 failed ---", fault, " ", l
				try:
					field, course1, course2, duration = l.split("-")
					course = course1+" "+course2
				except Exception, fault:
					print "try 2 failed ---", fault, " ", l
					raise


			course = course.strip()
			query = "select id from courses_course where name=\"%s\"" %(course)
			cursor.execute(query)
			data = cursor.fetchall()
			if len(data):
				#print "course found in listed - ", course
				# Adding college course mapping
				query = ("insert into colleges_college_courses (college_id, course_id) values"+
							"(%s,%s)"% (college_id, data[0][0]))
				#print query
				cursor.execute(query)
				data = cursor.fetchall()
				#print data
			else:
				#print "new course - ", course
				# creating new course with given details
				_b = ""
				entry = ("\""+course+"\",\""+_b+"\",\""+duration+"\",\""+_b+
					"\",\""+_b+"\",\""+_b+"\",\""+_b+ 
					"\",\""+_b+"\",\""+_b+"\",\""+_b+
					"\",\""+_b+"\",\""+_b+"\",\""+_b+
					"\",\""+_b+"\",\""+_b+"\",\""+_b+
					"\",\""+_b+"\",\""+_b+"\"")
				query = ("insert into courses_course (name, description, duration, level,"+
		 			"course_type, eligibility, overview, eligibility_2, syllabus,"+
		 			" suitability, beneficial, employment_areas, job_types,"+
		 			" advance_course, related_course, careers, recuiting_companies, other_details)"+
		 			" values ("+entry.encode('utf-8')+");")
				#print query
				cursor.execute(query)
				data = cursor.fetchall()
				#print data

				# mapping the new course with current college
				query = "select id from courses_course where name=\"%s\"" %(course)
				cursor.execute(query)
				data = cursor.fetchall()

				query = ("insert into colleges_college_courses (college_id, course_id) values"+
							"(%s,%s)"% (college_id, data[0][0]))
				cursor.execute(query)
				data = cursor.fetchall()

		db.commit()
		count += 1
	except Exception, fault:
		print "------------------------- "
		error_count += 1
		print "Could not parse row : "+str(index)+ " | Error : ", fault
		print "------------------------- "

print "Rows parsed : "+ str(count)
print "Rows errored : "+ str(error_count)


# Name of the Institution
# Address Line 1
# City
# State
# Zip
# Country
# Phone
# Mobile
# Fax
# Website
# Affiliated with	
# Approvals
# Courses